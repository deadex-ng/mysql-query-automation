"""Organisation Indicators Queries."""

from constants import FACILITIES, dates, start
import openpyxl
import string

path = "/home/buildr/Documents/PIH-projects/advanced data/mysql-query-automation/reports/data.xlsx"
wb = openpyxl.load_workbook(path)

def count_placeholders(str):
    return sum(1 for x in string.Formatter().parse(str) if x[1] is not None)

def write_header(program_row, gender_row, sheet, program, indicators):
    cur_sh = wb.get_sheet_by_name(sheet)
    cur_sh["A" + str(program_row)] = program
    cur_sh["B" + str(gender_row)] = "Male"
    cur_sh["C" + str(gender_row)] = "Female"
    i = gender_row + 1
    for indicator in indicators:
        cur_sh["A" + str(i)] = indicator
        i = i + 1
    wb.save(path)

def write_data(row, data):
    
    # dc = wb.get_sheet_by_name("Dambe Clinic")
    # ndh = wb.get_sheet_by_name("Lisungwi HC")
    dc = wb.get_sheet_by_name("Dambe Clinic")
    ndh = wb.get_sheet_by_name("Neno District Hospital")

    for x in data:
        # print(data)
        if x[0] == dc.title:
            print("Dambe")
            if x[1] == "M":
                dc["B" + str(row)] = x[2]
            if x[1] == "F":
                dc["C" + str(row)] = x[2]
        else:
            if x[1] == "M":
                value = ndh['B' + str(row)]
                value = value._value
                if isinstance(value, int):
                    value = value + x[2]
                    ndh["B" + str(row)] = value
                    wb.save(path)
                else:
                    ndh["B" + str(row)] = x[2]
                    wb.save(path)
                # value  = ndh.cell(row = row,column=2).value
                # if value is not None:
                #     value = value + x[2]
                #     ndh["B" + str(row)] = value
                # else:
                #     ndh["B" + str(row)] = x[2]
            if x[1] == "F":
                value = ndh['C' + str(row)]
                value = value._value
                if isinstance(value, int):
                    value = value + x[2]
                    ndh["C" + str(row)] = value
                    wb.save(path)
                else:
                    ndh["C" + str(row)] = x[2]
                    # print("NOW VALUE ON C10is ", ndh['C10']._value)
                    wb.save(path)
        wb.save(path)

# def write_data(row, data):
    
#     dc = wb.get_sheet_by_name("Dambe Clinic")
#     ndh = wb.get_sheet_by_name("Neno District Hospital")
#     for x in data:
#         # print(data)
#         if x[0] == dc.title:
#             print("Dambe")
#             if x[1] == "M":
#                 dc["B" + str(row)] = x[2]
#             if x[1] == "F":
#                 dc["C" + str(row)] = x[2]
#         else:
#             if x[1] == "M":
#                 value = ndh['B' + str(row)]
#                 value = value._value
#                 if isinstance(value, int):
#                     value = value + x[2]
#                     ndh["B" + str(row)] = value
#                     wb.save(path)
#                 else:
#                     ndh["B" + str(row)] = x[2]
#                     wb.save(path)
#                 # value  = ndh.cell(row = row,column=2).value
#                 # if value is not None:
#                 #     value = value + x[2]
#                 #     ndh["B" + str(row)] = value
#                 # else:
#                 #     ndh["B" + str(row)] = x[2]
#             if x[1] == "F":
#                 value = ndh['C' + str(row)]
#                 value = value._value
#                 if isinstance(value, int):
#                     value = value + x[2]
#                     ndh["C" + str(row)] = value
#                     wb.save(path)
#                 else:
#                     ndh["C" + str(row)] = x[2]
#                     # print("NOW VALUE ON C10is ", ndh['C10']._value)
#                     wb.save(path)
#         wb.save(path)


class CHF:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            1,
            2,
            "Dambe Clinic",
            "CHF",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with Diagonosis of rheumatic heart disease",
                "Patients with Diagonosis of congenital heart disease",
                "Patients with NYHA Classification recorded at last visit",
                "Patients recordd in NYHA class I",
                "Patients recordd in NYHA class II",
                "Patients recordd in NYHA class III",
                "Patients recordd in NYHA class IV",
                "Patients with a visit in last month, with a hospitalization in last month",
            ],
        )
        write_header(
            1,
            2,
            "Neno District Hospital",
            "CHF",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with Diagonosis of rheumatic heart disease",
                "Patients with Diagonosis of congenital heart disease",
                "Patients with NYHA Classification recorded at last visit",
                "Patients recordd in NYHA class I",
                "Patients recordd in NYHA class II",
                "Patients recordd in NYHA class III",
                "Patients recordd in NYHA class IV",
                "Patients with a visit in last month, with a hospitalization in last month",
            ],
        )

    def enrolled_and_active_in_care(self) -> None:
        end_date = '"2023-03-31"'
        end_reporting_date = '"2022-06-30"'
        query = """
        /*patients enrolled in active care*/
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---

        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select 

        lfo.location, gender, count(lfo.pat)

        from chronic_care_last_facility_outcome  lfo
        join mw_chf_initial chf
        on lfo.pat = chf.patient_id
        join mw_patient mp on mp.patient_id=chf.patient_id
        where state IN ("In advanced care")
        and pat in (
        SELECT patient_id FROM omrs_obs where encounter_type IN ("CHF_FOLLOWUP")   
        and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care: ", response)
        write_data(3, response)

        # for date in dates:
        #     formatted_qry = query.format(date)
        #     qry_lst = formatted_qry.split("---")
        #     for qry in qry_lst:
        #         self.cursor_obj.execute(qry)
        #     response = self.cursor_obj.fetchall()
        # write_data(3, response)

    def newly_registered(self):
        query = """
        /* Newly registered in reporting period */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate =  {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mci.patient_id)) as ncd_count
        from mw_chf_initial mci
        join mw_patient mp on mp.patient_id=mci.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mci.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(4, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_chf_initial where chf_initial_visit_id is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(5, response)

    def died(self):
        query = """
        /* Patients who died during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_chf_initial where chf_initial_visit_id is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(6, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* Patients with a visit in last month */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(7, response)

    def patients_with_diagnosis_rheumatic_heart_disease(self):
        query = """
        /* patients with diagnosis rheumatic heart disease */
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender,

        count(pat) as 'Rheumatic heart disease'

        from chronic_care_last_facility_outcome lfo
        join
        (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_chf_followup where visit_date <= @endDate
        group by patient_id
        ) patient_visit
        on patient_visit.patient_id = pat
        join mw_chf_initial mci
        on pat = mci.patient_id
        join mw_patient mp on mp.patient_id=mci.patient_id
        where pat IN (select patient_id from mw_chf_followup hf where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff and visit_date <= @endDate)
            and state IN ("In advanced Care") and diagnosis_rheumatic is not null
        group by lfo.location, gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_diagnosis_rheumatic_heart_disease: ", response)
        write_data(8, response)

    def patients_with_diagnosis_congenital_heart_disease(self):
        query = """
        /* patients with diagnosis Congenital heart disease */
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender,

        count(pat) as 'Congenital heart disease'


        from chronic_care_last_facility_outcome lfo
        join
        (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_chf_followup where visit_date <= @endDate
        group by patient_id 
        ) patient_visit
        on patient_visit.patient_id = pat
        join mw_chf_initial mci
        on pat = mci.patient_id
        join mw_patient mp on mp.patient_id=mci.patient_id
        where pat IN (select patient_id from mw_chf_followup hf where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff and visit_date <= @endDate)
            and state IN ("In advanced Care") and diagnosis_congenital is not null
        group by lfo.location, gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_diagnosis_congenital_heart_disease: ", response)
        write_data(9, response)

    def patients_with_nyha_classification_recorded_in_last_visit(self):
        query = """
        USE openmrs_warehouse;
        ---
        SET @startDate = '2021-04-01';
        ---
        SET @endDate = "2021-04-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")       
        and nyha_stage is not null
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_nyha_classification_recorded_in_last_visit: ", response)
        write_data(10, response)

    def patients_recordd_in_nyha_class_one(self):
        query = """
        /* Patients recorded in  NYHA class 1*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and nyha_stage="Nyha class 1"
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_recordd_in_nyha_class_one: ", response)
        write_data(11, response)

    def patients_recorded_in_nyha_class_two(self):
        query = """
        /* Patients recorded in  NYHA class 2*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and nyha_stage="Nyha class 2"
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_recorded_in_nyha_class_two: ", response)
        write_data(12, response)

    def patients_recorded_in_nyha_class_three(self):
        query = """
        /* Patients recorded in  NYHA class 3*/ 
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate); 
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and nyha_stage="Nyha class 3"
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_recorded_in_nyha_class_three: ", response)
        write_data(13, response)

    def patients_recorded_in_nyha_class_four(self):
        query = """
        /* Patients recorded in  NYHA class 4*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and nyha_stage="Nyha class 4"
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_recorded_in_nyha_class_four: ", response)
        write_data(14, response)

    def patients_with_a_visit_in_last_month_with_hospitilization_in_last_month(self):
        query = """
        /* Patients a visit in last month with hospitilization in last month*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select lfo.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_chf_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and hospitalized_since_last_visit_for_ncd="Yes"
        group by lfo.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_a_visit_in_last_month_with_hospitilization_in_last_month: ", response)        
        write_data(15, response)

class Epilepsy:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            17,
            18,
            "Dambe Clinic",
            "EPILEPSY",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with no seizures reported in last month",
                "Patients with a visit in last month with controlled seizure activity",
                "Patients with a visit in last month with hospitalization in the last month",
            ],
        )
        write_header(
            17,
            18,
            "Neno District Hospital",
            "EPILEPSY",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with no seizures reported in last month",
                "Patients with a visit in last month with controlled seizure activity",
                "Patients with a visit in last month with hospitalization in the last month",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /*patients enrolled in active care*/
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
 
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select
 
        lfo.location, gender, count(lfo.pat)
 
        from chronic_care_last_facility_outcome  lfo
        join mw_epilepsy_initial mei
        on lfo.pat = mei.patient_id
        join mw_patient mp on mp.patient_id=mei.patient_id
        where state IN ("In advanced care")
        and pat in (
        SELECT patient_id FROM omrs_obs where encounter_type IN ("Epilepsy_FOLLOWUP")   
        and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care",response) 
        write_data(19, response)

    def newly_registered(self):
        query = """
        /* 2. Newly registered in Epilepsy */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            lfo.location, gender, COUNT(pat)
        FROM
        mw_epilepsy_initial mei
        JOIN
        mw_patient mp ON mp.patient_id = mei.patient_id
        JOIN
        chronic_care_last_facility_outcome lfo ON lfo.pat = mei.patient_id
        WHERE
        visit_date BETWEEN @startDate AND @endDate
        AND state IN ('In advanced Care')
        GROUP BY location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(20, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
        ops.location,
        mp.gender,
        COUNT(DISTINCT (ops.patient_id)) AS ncd_count,
        'defaulted' AS defaulted
        FROM
        omrs_program_state ops
        JOIN
        omrs_patient_identifier opi ON ops.patient_id = opi.patient_id
        AND opi.type = 'Chronic Care Number'
        AND ops.location = opi.location
        JOIN
        mw_patient mp ON mp.patient_id = ops.patient_id
        WHERE
        ops.start_date BETWEEN @startDate AND @endDate
        AND state = 'patient defaulted'
        AND program = 'Chronic Care Program'
        AND ops.patient_id IN (SELECT 
            patient_id
        FROM
            mw_epilepsy_initial
        WHERE
            epilepsy_initial_visit_id IS NOT NULL)
        GROUP BY location 
        UNION SELECT 
        location,
        x.gender,
        COUNT(DISTINCT (patient_id)) AS ncd_count,
        'active but last appointment greater than cutoff' AS defaulted
        FROM
        (SELECT DISTINCT
        (mwp.patient_id),
            opi.identifier,
            mwp.first_name,
            mwp.last_name,
            ops.program,
            ops.state,
            ops.start_date,
            program_state_id,
            mwp.gender,
            ops.location,
            patient_visit.last_appt_date
        FROM
        mw_patient mwp
        JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
        JOIN (SELECT 
        patient_id, MAX(value_date) AS last_appt_date
        FROM
        omrs_obs
        WHERE
        encounter_type = 'EPILEPSY_FOLLOWUP'
            AND concept = 'Appointment date'
            AND obs_date <= @endDate
        GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
        JOIN (SELECT 
        index_desc,
            opi.patient_id AS pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
            start_date,
            program_state_id,
            end_date
        FROM
        (SELECT 
        @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:=patient_id
        FROM
        omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
        WHERE
        program IN ('Mental health care program')
            AND start_date <= @endDate
        ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
        JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
        AND opi.location = index_descending.location
        AND opi.type = 'Chronic Care Number'
        WHERE
        index_desc = 1) ops ON opi.patient_id = ops.pat
        AND opi.location = ops.location
        WHERE
        opi.type = 'Chronic Care Number'
            AND state IN ('In advanced Care')) x
        WHERE
        patient_id NOT IN (SELECT 
            patient_id
        FROM
            omrs_obs
        WHERE
            encounter_type = 'EPILEPSY_FOLLOWUP'
                AND concept = 'Appointment date'
                AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
        AND patient_id IN (SELECT 
            patient_id
        FROM
            omrs_obs
        WHERE
            encounter_type = 'EPILEPSY_FOLLOWUP'
                AND concept = 'Appointment date'
                AND value_date BETWEEN @startDate AND @endDate)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(21, response)

    def died(self):
        query = """
        /* Died*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
        ops.location,
        mp.gender,
        COUNT(DISTINCT (ops.patient_id)) AS ncd_count,
        'patient_died' AS patient_died  
        FROM
        omrs_program_state ops
        JOIN
        omrs_patient_identifier opi ON ops.patient_id = opi.patient_id
        AND opi.type = 'Chronic Care Number'
        AND ops.location = opi.location
        JOIN
        mw_patient mp ON mp.patient_id = ops.patient_id
        WHERE
        ops.start_date BETWEEN @startDate AND @endDate
        AND state = 'patient died'
        AND program = 'Chronic Care Program'
        AND ops.patient_id IN (SELECT 
            patient_id
        FROM
            mw_epilepsy_initial
        WHERE
            epilepsy_initial_visit_id IS NOT NULL)
        GROUP BY location 
        UNION SELECT 
        location,
        x.gender,
        COUNT(DISTINCT (patient_id)) AS ncd_count,
        'active but last appointment greater than cutoff' AS defaulted
        FROM
        (SELECT DISTINCT
        (mwp.patient_id),
            opi.identifier,
            mwp.first_name,
            mwp.last_name,
            ops.program,
            ops.state,
            ops.start_date,
            program_state_id,
            mwp.gender,
            ops.location,
            patient_visit.last_appt_date
        FROM
        mw_patient mwp
        JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
        JOIN (SELECT 
        patient_id, MAX(value_date) AS last_appt_date
        FROM
        omrs_obs
        WHERE
        encounter_type = 'EPILEPSY_FOLLOWUP'
            AND concept = 'Appointment date'
            AND obs_date <= @endDate
        GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
        JOIN (SELECT 
        index_desc,
            opi.patient_id AS pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
            start_date,
            program_state_id,
            end_date
        FROM
        (SELECT 
        @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:=patient_id
        FROM
        omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
        WHERE
        program IN ('Mental health care program')
            AND start_date <= @endDate
        ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
        JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
        AND opi.location = index_descending.location
        AND opi.type = 'Chronic Care Number'
        WHERE
        index_desc = 1) ops ON opi.patient_id = ops.pat
        AND opi.location = ops.location
        WHERE
        opi.type = 'Chronic Care Number'
            AND state IN ('In advanced Care')) x
        WHERE
        patient_id NOT IN (SELECT 
            patient_id
        FROM
            omrs_obs
        WHERE
            encounter_type = 'EPILEPSY_FOLLOWUP'
                AND concept = 'Appointment date'
                AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
        AND patient_id IN (SELECT 
            patient_id
        FROM
            omrs_obs
        WHERE
            encounter_type = 'EPILEPSY_FOLLOWUP'
                AND concept = 'Appointment date'
                AND value_date BETWEEN @startDate AND @endDate)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(22, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 3. Patients with a visit in last month*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
        mef.location, gender ,COUNT(DISTINCT (mef.patient_id))
        FROM
        mw_epilepsy_followup mef
        JOIN
        mw_patient mp ON mp.patient_id = mef.patient_id
        JOIN
        chronic_care_last_facility_outcome lfo ON lfo.pat = mef.patient_id
        WHERE
        visit_date BETWEEN @startDate AND @endDate
        AND state IN ('In advanced Care')
        GROUP BY mef.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(23, response)

    def patients_with_visit_last_month_with_no_seizures(self):
        query = """
        /* 4. Number of epilepsy patients with no seizures since last visit (in the last 3 months) */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
        mepf.location,
        gender,
        COUNT(DISTINCT (mepf.patient_id)) AS ncd_count
        FROM
        mw_epilepsy_followup mepf
        JOIN
        (SELECT 
        patient_id, MAX(visit_date) AS visit_date
        FROM
        mw_epilepsy_followup
        WHERE
        visit_date BETWEEN @startDate AND @endDate
        GROUP BY patient_id) mepf1 ON mepf1.patient_id = mepf.patient_id
        AND mepf.visit_date = mepf1.visit_date
        JOIN
        mw_patient mp ON mp.patient_id = mepf.patient_id
        JOIN
        chronic_care_last_facility_outcome lfo ON lfo.pat = mepf.patient_id
        WHERE
        (seizure_since_last_visit IS NULL
        OR seizure_since_last_visit = 'NO')
        AND state IN ('In advanced Care')
        GROUP BY location, gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_last_month_with_no_seizures: ", response)
        write_data(24, response)


class DiabetesType2:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            28,
            29,
            "Dambe Clinic",
            "Diabetes Type 2",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with an FBS < 150mg/Dl",
                "Currently enrolled patients that have ever experienced a complication",
                "Patients on Insulin",
            ],
        )
        write_header(
            28,
            29,
            "Neno District Hospital",
            "Diabetes Type 2",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with an FBS < 150mg/Dl",
                "Currently enrolled patients that have ever experienced a complication",
                "Patients on Insulin",
            ],
        )

    def enrolled_and_active_in_care(self) -> None:
        query = """
        /* Active in Care (Type 2) */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @endDate = {};
        ---
        select location, gender, count(*) as ncd_count
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_ncd_visits where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("On Treatment","In advanced Care")
            ) x
            where patient_id IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where (diagnosis_type_2_diabetes is not null))
            group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care:", response)
        write_data(30, response)

    def newly_registered(self):
        query = """
        /* 3. Newly registered Type 2 Diabetes */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhi.location, gender, count(*)
        from mw_diabetes_hypertension_initial dhi
        join mw_patient mp on mp.patient_id=dhi.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhi.patient_id
        where visit_date BETWEEN @startDate AND @endDate and (diagnosis_type_2_diabetes is not null)    
        and state IN ("In advanced Care")
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly registered: ", response)
        write_data(31, response)

    def defaulted(self):
        query = """
        /* Type 2 patients who have defaulted during the reporting period */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location,mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id NOT IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date BETWEEN @startDate AND @endDate)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
			group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(32, response)

    def died(self):
        query = """
        /* Type 2 patients who died during the reporting period */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location,mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id NOT IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date BETWEEN @startDate AND @endDate)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
			group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(33, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 5. Patients with a visit in last month (Type 2 Diabetes)*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhf.location, gender, count(distinct(dhf.patient_id)) as ncd_count
        from mw_diabetes_hypertension_followup dhf
        join mw_patient mp on mp.patient_id=dhf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and dhf.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
        and state IN ("In advanced Care")
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(34, response)

    def patients_with_visit_in_last_month_with_FBS_less_than_150(self):
        query = """
        /*7. Type 2 patients with a visit in the last month, with FBS <=150 mg/dL) */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhf.location, gender, count(distinct(dhf.patient_id)) as ncd_count
        from mw_diabetes_hypertension_followup dhf
        join mw_patient mp on mp.patient_id=dhf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and dhf.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_2_diabetes is not null)
        and fasting_blood_sugar <= 150
        and state IN ("In advanced Care")
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_last_month_with_FBS_less_than_150: ", response)
        write_data(35, response)

    def currently_enrolled_patients_that_have_ever_experienced_complication(self):
        query = """
        /* 10.Currently enrolled Type 2 patients that have ever experienced a complication */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select dhi.location,gender, count(distinct(dhi.patient_id)) as ncd_count
        from mw_diabetes_hypertension_initial dhi
        join mw_patient mp on mp.patient_id=dhi.patient_id
        where ((cardiovascular_disease is not null or retinopathy is not null or renal_disease is not null or stroke_and_tia is not null or
        peripheral_vascular_disease is not null or neuropathy is not null or sexual_disorder is not null) )
        and (diagnosis_type_2_diabetes is not null)
        and dhi.patient_id in
        (   
	    select distinct(patient_id)
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where (diagnosis_type_2_diabetes is not null))
        )
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("currently_enrolled_patients_that_have_ever_experienced_complication",
            response,
        )
        write_data(36, response)

    def patients_on_insulin(self):
        query = """
        /*8.Type 2 patients on  Insulin */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SET @defaultCutOff = 60;
        ---
        select location, gender, count(distinct(dhf.patient_id)) as ncd_count
        from mw_diabetes_hypertension_followup dhf
        join (
	    select patient_id, max(visit_date) as last_visit_date
        from mw_diabetes_hypertension_followup where visit_date <=  @endDate
        group by patient_id
        ) dhf1
        on dhf.patient_id = dhf1.patient_id and dhf.visit_date = dhf1.last_visit_date
        join mw_patient mp on mp.patient_id=dhf.patient_id
        where (dhf.diabetes_med_long_acting is not null or dhf.diabetes_med_short_acting is not null)
        and dhf.patient_id in (
        select patient_id
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_ncd_visits where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date    
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date >=@defaultCutOff)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where (diagnosis_type_2_diabetes is not null))
            )    
            group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_on_insulin", response)
        write_data(37, response)


class DiabetesType1:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            39,
            40,
            "Dambe Clinic",
            "Diabetes Type 1",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with an FBS < 150mg/Dl",
            ],
        )
        write_header(
            39,
            40,
            "Neno District Hospital",
            "Diabetes Type 1",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month with an FBS < 150mg/Dl",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /* Active in Care (Type 1)*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @endDate = {};
        ---
        select location, gender, count(*) as ncd_count
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_ncd_visits where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where (diagnosis_type_1_diabetes is not null))
            group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care", response)
        write_data(41, response)

    def newly_registered(self):
        query = """
        /* 2. Newly registered Type 1 Diabetes */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhi.location, gender, count(*)
        from mw_diabetes_hypertension_initial dhi
        join mw_patient mp on mp.patient_id=dhi.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhi.patient_id
        where visit_date BETWEEN @startDate AND @endDate and (diagnosis_type_1_diabetes is not null)
        and state IN ("In advanced Care")
        group by dhi.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered", response)
        write_data(42, response)

    def defaulted(self):
        query = """
        /* Type 1 patients who have defaulted during the reporting period */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = "2021-04-01";
        ---
        SET @endDate = "2021-06-30";
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
        group by location
        union 
        select location,x.gender,count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id NOT IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date BETWEEN @startDate AND @endDate)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
			group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(43, response)

    def died(self):
        query = """
        /* Type 1 patients who have died during the reporting period */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
        group by location
        union 
        select location,x.gender,count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
            ) patient_visit
            on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
            opi.patient_id as pat,
            opi.identifier,
            index_descending.state,
            index_descending.location,
            index_descending.program,
        start_date,
            program_state_id,
            end_date
        FROM (SELECT
            @r:= IF(@u = patient_id, @r + 1,1) index_desc,
            location,
            state,
            program,
            start_date,
            end_date,
            patient_id,
            program_state_id,
            @u:= patient_id
        FROM omrs_program_state,
                    (SELECT @r:= 1) AS r,
                    (SELECT @u:= 0) AS u
                    where program IN ("Chronic Care Program")
        and start_date <= @endDate
            ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
            ) index_descending
            join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
            and opi.location = index_descending.location
            and opi.type = "Chronic Care Number"
            where index_desc = 1)
            ops
            on opi.patient_id = ops.pat and opi.location = ops.location
            where opi.type = "Chronic Care Number"
            and state IN ("In advanced Care")
            ) x
            where patient_id NOT IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
            and patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date BETWEEN @startDate AND @endDate)
            and patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
			group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(44, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 4. Patients with a visit in last month (Type 1 Diabetes)*/
        USE openmrs_warehouse;
        ---
        SET @startDate = '2021-04-01';
        ---
        SET @endDate = "2021-04-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhf.location, gender, count(distinct(dhf.patient_id)) as ncd_count
        from mw_diabetes_hypertension_followup dhf
        join mw_patient mp on mp.patient_id=dhf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and dhf.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
        and state IN ("In advanced Care")
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(45, response)

    def patients_with_visit_in_last_month_with_FBS_less_than_150(self):
        query = """
        /*6. Type 1 patients with a visit in the last month, with FBS <=150 mg/dL) */
        USE openmrs_warehouse;
        ---
        SET @startDate = '2021-04-01';
        ---
        SET @endDate = "2021-06-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select dhf.location, gender, count(distinct(dhf.patient_id)) as ncd_count
        from mw_diabetes_hypertension_followup dhf  
        join mw_patient mp on mp.patient_id=dhf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=dhf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and dhf.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_type_1_diabetes is not null)
        and fasting_blood_sugar <= 150
        and state IN ("In advanced Care")
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_last_month_with_FBS_less_than_150: ", response)
        write_data(46, response)


class CKD:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            48,
            49,
            "Dambe Clinic",
            "CKD",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month creatinine level recorded in last month",
                "Patients with a visit in last month with urinalysis recorded in the last month",
                "Patients with a visit in last month with hospitalization in the last month",
            ],
        )
        write_header(
            48,
            49,
            "Neno District Hospital",
            "CKD",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in last month creatinine level recorded in last month",
                "Patients with a visit in last month with urinalysis recorded in the last month",
                "Patients with a visit in last month with hospitalization in the last month",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /*patients enrolled in active care*/
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select 

        ckd.location, gender, count(lfo.pat)

        from chronic_care_last_facility_outcome  lfo
        join mw_ckd_initial ckd
        on lfo.pat = ckd.patient_id
        join mw_patient mp on mp.patient_id=ckd.patient_id
        where state IN ("In advanced care")
        and pat in (
        SELECT patient_id FROM omrs_obs where encounter_type IN ("CKD_FOLLOWUP")   
        and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        group by ckd.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care: ", response)
        write_data(50, response)

    def newly_registered(self):
        query = """
        /* Newly registered in reporting period */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  mci.location, gender, count(distinct(mci.patient_id)) as ncd_count
        from mw_ckd_initial mci
        join mw_patient mp on mp.patient_id=mci.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mci.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        group by mci.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(51, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_ckd_initial where ckd_initial_visit_id is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "CKD_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "CKD_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(52, response)

    def died(self):
        query = """
        /* Patients who died during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_ckd_initial where ckd_initial_visit_id is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "CHF_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,   
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "CKD_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "CKD_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(53, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* Patients with a visit in last month */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select mcf.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from  mw_ckd_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        group by mcf.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(54, response)

    def patients_with_visit_in_last_month_creatinine_level_recorded(self):
        query = """
        /* Patients a visit in last month with cretinine level recorded in last month*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  mcf.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from   mw_ckd_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and creatinine is not null
        group by mcf.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_last_month_creatinine_level_recorded: ", response)
        write_data(55, response)

    def patients_with_visit_in_last_month_with_urinalysis_recorded(self):
        query = """
        /* Patients a visit in last month with urinalysis recorded in last month*/
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  mcf.location, gender, count(distinct(mcf.patient_id)) as ncd_count
        from   mw_ckd_followup mcf
        join mw_patient mp on mp.patient_id=mcf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mcf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and state IN ("In advanced Care")
        and urine_protein is not null
        group by mcf.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_last_month_with_urinalysis_recorded: ", response)
        write_data(56, response)

class Mentalhealth:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            59,
            60,
            "Dambe Clinic",
            "Mental Health",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in the last month who have been hospitalized in the last month",
                "Patients on medicine who reported side effects from medication (In last month)",
                "Patients in care that were reported stable in last visit (last month)",
            ],
        )
        write_header(
            59,
            60,
            "Neno District Hospital",
            "Mental Health",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Patients with a visit in the last month who have been hospitalized in the last month",
                "Patients on medicine who reported side effects from medication (In last month)",
                "Patients in care that were reported stable in last visit (last month)",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /*patients enrolled in active care*/
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
 
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        SELECT 
        lfo.location, gender, COUNT(lfo.pat)
        FROM
        chronic_care_last_facility_outcome lfo
        JOIN
        mw_mental_health_initial mmhi ON lfo.pat = mmhi.patient_id
        JOIN
        mw_patient mp ON mp.patient_id = mmhi.patient_id
        WHERE
        state IN ('In advanced care')
        AND pat IN (SELECT 
            patient_id
        FROM
            omrs_obs
        WHERE
            encounter_type IN ('MENTAL_HEALTH_FOLLOWUP')
                AND concept = 'Appointment date'
                AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care: ", response)
        write_data(61, response)

    def newly_registered(self):
        query = """
        /* 2. Newly registered in Mental Health */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mmhi.location,mp.gender,COUNT(*) 
        FROM
        mw_mental_health_initial mmhi
        JOIN
        mw_patient mp ON mp.patient_id = mmhi.patient_id
        JOIN
        chronic_care_last_facility_outcome lfo ON lfo.pat = mmhi.patient_id
        WHERE
        visit_date BETWEEN @startDate AND @endDate
        AND state IN ('In advanced Care')
        GROUP BY mmhi.location,mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(62, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
            ops.location,
            mp.gender,
            COUNT(DISTINCT (ops.patient_id)) AS ncd_count,
            'defaulted' AS defaulted
        FROM
            omrs_program_state ops
                JOIN
            omrs_patient_identifier opi ON ops.patient_id = opi.patient_id
                AND opi.type = 'Chronic Care Number'
                AND ops.location = opi.location
                JOIN
            mw_patient mp ON mp.patient_id = ops.patient_id
        WHERE
            ops.start_date BETWEEN @startDate AND @endDate
                AND state = 'patient defaulted'
                AND program = 'Chronic Care Program'
                AND ops.patient_id IN (SELECT 
                    patient_id
                FROM
                    mw_mental_health_initial
                WHERE
                    mental_health_initial_visit_id IS NOT NULL)
        GROUP BY location 
        UNION SELECT 
            location,
            x.gender,
            COUNT(DISTINCT (patient_id)) AS ncd_count,
            'active but last appointment greater than cutoff' AS defaulted
        FROM
            (SELECT DISTINCT
                (mwp.patient_id),
                    opi.identifier,
                    mwp.first_name,
                    mwp.last_name,
                    ops.program,
                    ops.state,
                    ops.start_date,
                    program_state_id,
                    mwp.gender,
                    ops.location,
                    patient_visit.last_appt_date
            FROM
                mw_patient mwp
            JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
            JOIN (SELECT 
                patient_id, MAX(value_date) AS last_appt_date
            FROM
                omrs_obs
            WHERE
                encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                    AND concept = 'Appointment date'
                    AND obs_date <= @endDate
            GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
            JOIN (SELECT 
                index_desc,
                    opi.patient_id AS pat,
                    opi.identifier,
                    index_descending.state,
                    index_descending.location,
                    index_descending.program,
                    start_date,
                    program_state_id,
                    end_date
            FROM
                (SELECT 
                @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
                    location,
                    state,
                    program,
                    start_date,
                    end_date,
                    patient_id,
                    program_state_id,
                    @u:=patient_id
            FROM
                omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
            WHERE
                program IN ('Mental health care program')
                    AND start_date <= @endDate
            ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
            JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
                AND opi.location = index_descending.location
                AND opi.type = 'Chronic Care Number'
            WHERE
                index_desc = 1) ops ON opi.patient_id = ops.pat
                AND opi.location = ops.location
            WHERE
                opi.type = 'Chronic Care Number'
                    AND state IN ('In advanced Care')) x
        WHERE
            patient_id NOT IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
                AND patient_id IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND value_date BETWEEN @startDate AND @endDate)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(63, response)

    def died(self):
        query = """
        /* Patients who died during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
            ops.location,
            mp.gender,
            COUNT(DISTINCT (ops.patient_id)) AS ncd_count,
            'died' AS died
        FROM
            omrs_program_state ops
                JOIN
            omrs_patient_identifier opi ON ops.patient_id = opi.patient_id
                AND opi.type = 'Chronic Care Number'
                AND ops.location = opi.location
                JOIN
            mw_patient mp ON mp.patient_id = ops.patient_id
        WHERE
            ops.start_date BETWEEN @startDate AND @endDate
                AND state = 'patient died'
                AND program = 'Chronic Care Program'
                AND ops.patient_id IN (SELECT 
                    patient_id
                FROM
                    mw_mental_health_initial
                WHERE
                    mental_health_initial_visit_id IS NOT NULL)
        GROUP BY location 
        UNION SELECT 
            location,
            x.gender,
            COUNT(DISTINCT (patient_id)) AS ncd_count,
            'died' AS died
        FROM
            (SELECT DISTINCT
                (mwp.patient_id),
                    opi.identifier,
                    mwp.first_name,
                    mwp.last_name,
                    ops.program,
                    ops.state,
                    ops.start_date,
                    program_state_id,
                    mwp.gender,
                    ops.location,
                    patient_visit.last_appt_date
            FROM
                mw_patient mwp
            JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
            JOIN (SELECT 
                patient_id, MAX(value_date) AS last_appt_date
            FROM
                omrs_obs
            WHERE
                encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                    AND concept = 'Appointment date'
                    AND obs_date <= @endDate
            GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
            JOIN (SELECT 
                index_desc,
                    opi.patient_id AS pat,
                    opi.identifier,
                    index_descending.state,
                    index_descending.location,
                    index_descending.program,
                    start_date,
                    program_state_id,
                    end_date
            FROM
                (SELECT 
                @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
                    location,
                    state,
                    program,
                    start_date,
                    end_date,
                    patient_id,
                    program_state_id,
                    @u:=patient_id
            FROM
                omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
            WHERE
                program IN ('Mental health care program')
                    AND start_date <= @endDate
            ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
            JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
                AND opi.location = index_descending.location
                AND opi.type = 'Chronic Care Number'
            WHERE
                index_desc = 1) ops ON opi.patient_id = ops.pat
                AND opi.location = ops.location
            WHERE
                opi.type = 'Chronic Care Number'
                    AND state IN ('In advanced Care')) x
        WHERE
            patient_id NOT IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
                AND patient_id IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'MENTAL_HEALTH_FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND value_date BETWEEN @startDate AND @endDate)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(64, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 3. Patients with a visit in last month */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mmhf.location, gender, COUNT(DISTINCT (mmhf.patient_id))
        FROM
            mw_mental_health_followup mmhf
                JOIN
            mw_patient mp ON mp.patient_id = mmhf.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mmhf.patient_id
        WHERE
            visit_date BETWEEN @startDate AND @endDate
                AND state IN ('In advanced Care')
        GROUP BY mmhf.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(65, response)

    def patients_with_visit_in_last_month_who_have_been_hospitalized(self):
        query = """
        /* 6. Patients with a visit in the last month who have been hospitalized in the last month */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            
            gender,
            COUNT(DISTINCT (mmhf.patient_id)) AS ncd_count
        FROM
            mw_mental_health_followup mmhf
                JOIN
            (SELECT 
                patient_id, MAX(visit_date) AS visit_date
            FROM
                mw_mental_health_followup
            WHERE
                visit_date BETWEEN @startDate AND @endDate
            GROUP BY patient_id) mmhf1 ON mmhf1.patient_id = mmhf.patient_id
                AND mmhf.visit_date = mmhf1.visit_date
                JOIN
            mw_patient mp ON mp.patient_id = mmhf1.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mmhf1.patient_id
        WHERE
        state IN ('In advanced Care') and
            hospitalized_since_last_visit = 'Yes'
        GROUP BY gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_with_visit_in_last_month_who_have_been_hospitalized: ", response
        )
        write_data(66, response)

    def patients_on_medicine_who_reported_side_effects_from_medication(self):
        query = """
        /* 4. Patients on medication who reported side effects at the last visit (in the last month)  */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mmhf.location,
            gender,
            COUNT(DISTINCT (mmhf.patient_id)) AS ncd_count
        FROM
            mw_mental_health_followup mmhf
                JOIN
            (SELECT 
                patient_id, MAX(visit_date) AS visit_date
            FROM
                mw_mental_health_followup
            WHERE
                visit_date BETWEEN @startDate AND @endDate
            GROUP BY patient_id) mmhf1 ON mmhf1.patient_id = mmhf.patient_id
                AND mmhf.visit_date = mmhf1.visit_date
                JOIN
            mw_patient mp ON mp.patient_id = mmhf1.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mmhf1.patient_id
        WHERE
            medications_side_effects = TRUE
                AND state IN ('In advanced Care')
        group by location,gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_on_medicine_who_reported_side_effects_from_medication: ", response
        )
        write_data(67, response)

    def patients_in_care_that_were_reported_stable_in_last_visit(self):
        query = """
        /* 5. Patients in care that were reported as stable at last visit (in the last month)  */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mmhf.location,
            gender,
            COUNT(DISTINCT (mmhf.patient_id)) AS ncd_count
        FROM
            mw_mental_health_followup mmhf
                JOIN
            (SELECT 
                patient_id, MAX(visit_date) AS visit_date
            FROM
                mw_mental_health_followup
            WHERE
                visit_date BETWEEN @startDate AND @endDate
            GROUP BY patient_id) mmhf1 ON mmhf1.patient_id = mmhf.patient_id
                AND mmhf.visit_date = mmhf1.visit_date
                JOIN
            mw_patient mp ON mp.patient_id = mmhf1.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mmhf1.patient_id
        WHERE
            patient_stable = 'yes'
                AND state IN ('In advanced Care')
        GROUP BY location,gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_on_medicine_who_reported_side_effects_from_medication: ", response
        )
        write_data(68, response)


class COPD:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            70,
            71,
            "Dambe Clinic",
            "COPD",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
            ],
        )
        write_header(
            70,
            71,
            "Neno District Hospital",
            "COPD",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /* Patients enrolled and active in care1*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select location,gender, count(distinct(patient_id)) as ncd_count
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_copd is not null)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care: ", response)
        write_data(72, response)

    def newly_registered(self):
        query = """
        /* 1. Newly registered */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select mai.location, gender, count(distinct(mai.patient_id)) as ncd_count
        from mw_asthma_initial mai
        join mw_patient mp on mp.patient_id=mai.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mai.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and diagnosis_copd is not null
        and state IN ("In advanced Care")
        group by location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(73, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_asthma_initial where diagnosis_copd is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_copd is not null)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(74, response)

    def died(self):
        query = """
        /* Patients who have died during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_asthma_initial where diagnosis_copd is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "died" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_copd is not null)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(75, response)


class Asthma:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            77,
            78,
            "Dambe Clinic",
            "Asthma",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in last month",
                "Patients with disease severity recorded at most recent visit",
                "Patients with Intermittent asthma recorded at recent visit",
                "Patients with mild persitent asthma recorded at recent visit",
                "Patients with moderate asthma recorded at recent visit",
                "Patients with severe persitent asthma recorded at recent visit",
                "Patients with uncontrolled asthma recorded at recent visit",
            ],
        )
        write_header(
            77,
            78,
            "Neno District Hospital",
            "Asthma",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in last month",
                "Patients with disease severity recorded at most recent visit",
                "Patients with Intermittent asthma recorded at recent visit",
                "Patients with mild persitent asthma recorded at recent visit",
                "Patients with moderate asthma recorded at recent visit",
                "Patients with severe persitent asthma recorded at recent visit",
                "Patients with uncontrolled asthma recorded at recent visit",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /* Patients enrolled and active in care1*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select x.location, gender,count(distinct(patient_id)) as ncd_count
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        group by  x.location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care: ", response)
        write_data(79, response)

    def newly_registered(self):
        query = """
        /* 1. Newly registered */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select mai.location, gender, count(distinct(mai.patient_id)) as ncd_count
        from mw_asthma_initial mai
        join mw_patient mp on mp.patient_id=mai.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=mai.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and diagnosis_asthma is not null
        and state IN ("In advanced Care")
        group by mai.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered: ", response)
        write_data(80, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "defaulted" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient defaulted" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "active but last appointment greater than cutoff" as defaulted
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted: ", response)
        write_data(81, response)

    def died(self):
        query = """
        /* Patients who have died during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as died
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        group by location
        union 
        select location,x.gender, count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        SELECT patient_id, MAX(value_date) as last_appt_date FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and obs_date <= @endDate
        group by patient_id
        ) patient_visit on patient_visit.patient_id = mwp.patient_id
        JOIN
        (SELECT
        index_desc,
        opi.patient_id as pat,
        opi.identifier,
        index_descending.state,
        index_descending.location,
        index_descending.program,
        start_date,
        program_state_id,
        end_date
        FROM (SELECT
        @r:= IF(@u = patient_id, @r + 1,1) index_desc,
        location,
        state,
        program,
        start_date,
        end_date,
        patient_id,
        program_state_id,
        @u:= patient_id
        FROM omrs_program_state,
        (SELECT @r:= 1) AS r,
        (SELECT @u:= 0) AS u
        where program IN ("Chronic Care Program")
        and start_date <= @endDate
        ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
        ) index_descending
        join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
        and opi.location = index_descending.location
        and opi.type = "Chronic Care Number"
        where index_desc = 1)
        ops
        on opi.patient_id = ops.pat and opi.location = ops.location
        where opi.type = "Chronic Care Number"
        and state IN ("In advanced Care")
        ) x
        where patient_id NOT IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and floor(datediff(@endDate,value_date)) <=  @defaultCutOff)
        and patient_id IN (SELECT patient_id FROM omrs_obs where encounter_type = "ASTHMA_FOLLOWUP" and concept = "Appointment date" and value_date BETWEEN @startDate AND @endDate)
        and patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died: ", response)
        write_data(82, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 2. Patients with a visit in last months */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  maf.location, gender, count(distinct(maf.patient_id)) as ncd_count
        from mw_asthma_followup maf
        join mw_patient mp on mp.patient_id=maf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=maf.patient_id
        where visit_date BETWEEN @startDate AND @endDate
        and maf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
        and state IN ("In advanced Care")
        group by maf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month: ", response)
        write_data(83, response)

    def patients_with_disease_severity_recorded_at_most_recent_visit(self):
        query = """
        /* 3. Patients with disease severity recorded at most recent visit  */
        use openmrs_warehouse;
        ---
        SET @startDate = "2021-04-01";
        ---
        SET @endDate = "2021-06-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select   asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity is not null
        and state IN ("In advanced Care")
        group by  asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_with_disease_severity_recorded_at_most_recent_visit: ", response
        )
        write_data(84, response)

    def patients_with_intermittent_asthma_recorded_at_recent_visit(self):
        query = """
        /* 4. Patients with disease controlled (severity intermittent at last visit) */
        use openmrs_warehouse;
        ---
        SET @startDate = "2021-04-01";
        ---
        SET @endDate = "2021-06-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity IN ("Intermittent")
        and state IN ("In advanced Care")
        group by  asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_intermittent_asthma_recorded_at_recent_visit: ", response)
        write_data(85, response)

    def patients_with_mild_persitent_asthma_recorded_at_recent_visit(self):
        query = """
        /* 5. Patients with disease controlled (severity Mild persistent at last visit) */
        use openmrs_warehouse;
        ---
        SET @startDate = "2021-04-01";
        ---
        SET @endDate = "2021-06-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity IN ("Mild persistent")
        and state IN ("In advanced Care")
        group by  asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_with_mild_persitent_asthma_recorded_at_recent_visit: ", response
        )
        write_data(86, response)

    def patients_with_moderate_asthma_recorded_at_recent_visit(self):
        query = """
        /* 7. Patients with disease controlled (severity Moderate persistent at last visit) */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity IN ("Moderate persistent")
        and state IN ("In advanced Care")
        group by asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_moderate_asthma_recorded_at_recent_visit: ", response)
        write_data(87, response)

    def patients_with_severe_persitent_asthma_recorded_at_recent_visit(self):
        query = """
        /* 6. Patients with disease controlled (severity Severe persistent at last visit) */
        use openmrs_warehouse;
        ---
        SET @startDate = "2021-04-01";
        ---
        SET @endDate = "2021-06-30";
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity IN ("Severe persistent")
        and state IN ("In advanced Care")
        group by  asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "patients_with_severe_persitent_asthma_recorded_at_recent_visit:", response
        )
        write_data(88, response)

    def patients_with_uncontrolled_asthma_recorded_at_recent_visit(self):
        query = """
        /* 8. Patients with disease uncontrolled */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        select  asf.location, gender, count(distinct(asf.patient_id)) as ncd_count
        from mw_asthma_followup asf
        join mw_patient mp on mp.patient_id=asf.patient_id
        join chronic_care_last_facility_outcome lfo on lfo.pat=asf.patient_id
        join (
            select patient_id,max(visit_date) as visit_date from mw_asthma_followup 
            where visit_date BETWEEN @startDate AND @endDate
            group by patient_id
            ) asf1
            on asf1.patient_id = asf.patient_id and asf.visit_date = asf1.visit_date
            and asf.patient_id in (select patient_id from mw_asthma_initial where diagnosis_asthma is not null)
            and asthma_severity IN ("Severe uncontrolled")
        and state IN ("In advanced Care")
        group by asf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_uncontrolled_asthma_recorded_at_recent_visit:", response)
        write_data(89, response)


class Hypertension:
    def __init__(self, cursor_obj):
        self.cursor_obj = cursor_obj
        write_header(
            91,
            92,
            "Dambe Clinic",
            "Hypertension",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Currently enrolled patients that have ever experienced a complication",
                "Patients with a visit in last 3 months (excluding new patients) that have BPbelow 140/90",
            ],
        )
        write_header(
            91,
            92,
            "Neno District Hospital",
            "Hypertension",
            [
                "Enrolled and active in care",
                "Newly registered in reporting period",
                "Patients who have defaulted during the reporting period",
                "Patients who died during the reporting period",
                "Patients with a visit in the last month",
                "Currently enrolled patients that have ever experienced a complication",
                "Patients with a visit in last 3 months (excluding new patients) that have BPbelow 140/90",
            ],
        )

    def enrolled_and_active_in_care(self):
        query = """
        /*patients enrolled in active care*/
        use openmrs_warehouse;
        ---
        SET @endDate = {};
        ---
        
        call create_chronic_care_outcome(@endDate);
        ---
        SET @defaultCutOff = 60;
        ---
        SELECT 
            lfo.location, gender, COUNT(lfo.pat)
        FROM
            chronic_care_last_facility_outcome lfo
                JOIN
            mw_diabetes_hypertension_initial mdhi ON lfo.pat = mdhi.patient_id
                JOIN
            mw_patient mp ON mp.patient_id = mdhi.patient_id
        WHERE
            state IN ('In advanced care')
                AND pat IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type IN ('DIABETES HYPERTENSION FOLLOWUP')
                        AND concept = 'Appointment date'
                        AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("enrolled_and_active_in_care:", response)
        write_data(93, response)

    def newly_registered(self):
        query = """
        /* 2. Newly registered Hypertension */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mdhi.location, gender, COUNT(*)
        FROM
            mw_diabetes_hypertension_initial mdhi
                JOIN
            mw_patient mp ON mp.patient_id = mdhi.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mdhi.patient_id
        WHERE
            visit_date BETWEEN @startDate AND @endDate
                AND state IN ('In advanced Care')
                AND diagnosis_hypertension IS NOT NULL
        GROUP BY mdhi.location , mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("newly_registered:", response)
        write_data(94, response)

    def defaulted(self):
        query = """
        /* Patients who have defaulted during the reporting period*/
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
            ops.location,
            mp.gender,
            COUNT(DISTINCT (ops.patient_id)) AS ncd_count,
            'defaulted' AS defaulted
        FROM
            omrs_program_state ops
                JOIN
            omrs_patient_identifier opi ON ops.patient_id = opi.patient_id
                AND opi.type = 'Chronic Care Number'
                AND ops.location = opi.location
                JOIN
            mw_patient mp ON mp.patient_id = ops.patient_id
        WHERE
            ops.start_date BETWEEN @startDate AND @endDate
                AND state = 'patient defaulted'
                AND program = 'Chronic Care Program'
                AND ops.patient_id IN (SELECT 
                    patient_id
                FROM
                    mw_diabetes_hypertension_initial
                WHERE
                    initial_visit_id IS NOT NULL)
        GROUP BY location 
        UNION SELECT 
            location,
            x.gender,
            COUNT(DISTINCT (patient_id)) AS ncd_count,
            'active but last appointment greater than cutoff' AS defaulted
        FROM
            (SELECT DISTINCT
                (mwp.patient_id),
                    opi.identifier,
                    mwp.first_name,
                    mwp.last_name,
                    ops.program,
                    ops.state,
                    ops.start_date,
                    program_state_id,
                    mwp.gender,
                    ops.location,
                    patient_visit.last_appt_date
            FROM
                mw_patient mwp
            JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
            JOIN (SELECT 
                patient_id, MAX(value_date) AS last_appt_date
            FROM
                omrs_obs
            WHERE
                encounter_type = 'DIABETES HYPERTENSION FOLLOWUP'
                    AND concept = 'Appointment date'
                    AND obs_date <= @endDate
            GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
            JOIN (SELECT 
                index_desc,
                    opi.patient_id AS pat,
                    opi.identifier,
                    index_descending.state,
                    index_descending.location,
                    index_descending.program,
                    start_date,
                    program_state_id,
                    end_date
            FROM
                (SELECT 
                @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
                    location,
                    state,
                    program,
                    start_date,
                    end_date,
                    patient_id,
                    program_state_id,
                    @u:=patient_id
            FROM
                omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
            WHERE
                program IN ('Chronic Care Program')
                    AND start_date <= @endDate
            ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
            JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
                AND opi.location = index_descending.location
                AND opi.type = 'Chronic Care Number'
            WHERE
                index_desc = 1) ops ON opi.patient_id = ops.pat
                AND opi.location = ops.location
            WHERE
                opi.type = 'Chronic Care Number'
                    AND state IN ('In advanced Care')) x
        WHERE
            patient_id NOT IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'DIABETES HYPERTENSION FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND FLOOR(DATEDIFF(@endDate, value_date)) <= @defaultCutOff)
                AND patient_id IN (SELECT 
                    patient_id
                FROM
                    omrs_obs
                WHERE
                    encounter_type = 'DIABETES HYPERTENSION FOLLOWUP'
                        AND concept = 'Appointment date'
                        AND value_date BETWEEN @startDate AND @endDate)
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("defaulted:", response)
        write_data(95, response)

    def died(self):
        query = """
        /* patients who have died during the reporting period */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = 60;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        select ops.location, mp.gender, count(distinct(ops.patient_id)) as ncd_count, "died" as defaulted
        from omrs_program_state ops
        join omrs_patient_identifier opi
        on ops.patient_id = opi.patient_id and opi.type = "Chronic Care Number"
        and ops.location = opi.location
        join mw_patient mp on mp.patient_id=ops.patient_id
        where ops.start_date between @startDate and @endDate
        and state = "patient died" and program = "Chronic Care Program"
        and ops.patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_hypertension is not null)
        group by location
        union 
        select location,x.gender,count(distinct(patient_id)) as ncd_count, "died" as died
        from
        (
        select distinct(mwp.patient_id), opi.identifier, mwp.first_name, mwp.last_name, ops.program, ops.state,ops.start_date,program_state_id,  mwp.gender,
        ops.location, patient_visit.last_appt_date
        from  mw_patient mwp  
        join omrs_patient_identifier opi
        on mwp.patient_id = opi.patient_id
        join (
        select patient_id,MAX(visit_date) as visit_date ,MAX(next_appointment_date) as last_appt_date from mw_diabetes_hypertension_followup where visit_date <= @endDate
        group by patient_id
                    ) patient_visit
                    on patient_visit.patient_id = mwp.patient_id
        JOIN
                (SELECT
        index_desc,
                    opi.patient_id as pat,
                    opi.identifier,
                    index_descending.state,
                    index_descending.location,
                    index_descending.program,
        start_date,
                    program_state_id,
                    end_date
        FROM (SELECT
                    @r:= IF(@u = patient_id, @r + 1,1) index_desc,
                    location,
                    state,
                    program,
                    start_date,
                    end_date,
                    patient_id,
                    program_state_id,
                    @u:= patient_id
            FROM omrs_program_state,
                            (SELECT @r:= 1) AS r,
                            (SELECT @u:= 0) AS u
                            where program IN ("Chronic Care Program")
        and start_date <= @endDate
                    ORDER BY patient_id DESC, start_date DESC, program_state_id DESC
                    ) index_descending
                    join omrs_patient_identifier opi on index_descending.patient_id = opi.patient_id
                    and opi.location = index_descending.location
                    and opi.type = "Chronic Care Number"
                    where index_desc = 1)
                    ops
                    on opi.patient_id = ops.pat and opi.location = ops.location
                    where opi.type = "Chronic Care Number"
                    and state IN ("In advanced Care")
                    ) x
                    where patient_id NOT IN (select patient_id from mw_diabetes_hypertension_followup where floor(datediff(@endDate,next_appointment_date)) <=  @defaultCutOff)
                    and patient_id IN (select patient_id from mw_diabetes_hypertension_followup where next_appointment_date BETWEEN @startDate AND @endDate)
                    and patient_id in (select patient_id from mw_diabetes_hypertension_initial where diagnosis_hypertension is not null)
                    group by location, gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("died:", response)
        write_data(96, response)

    def patients_with_visit_in_the_last_month(self):
        query = """
        /* 3. Patients with a visit in last month */
        USE openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            mdhf.location,
            gender,
            COUNT(DISTINCT (mdhf.patient_id)) AS ncd_count
        FROM
            mw_diabetes_hypertension_followup mdhf
                JOIN
            mw_patient mp ON mp.patient_id = mdhf.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = mdhf.patient_id
        WHERE
            visit_date BETWEEN @startDate AND @endDate
                AND state IN ('In advanced Care')
                AND mdhf.patient_id IN (SELECT 
                    patient_id
                FROM
                    mw_diabetes_hypertension_initial
                WHERE
                    diagnosis_hypertension IS NOT NULL)
        GROUP BY mdhf.location, gender
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_the_last_month:", response)
        write_data(97, response)

    def currently_enrolled_patients_that_have_ever_experienced_complication(self):
        query = """
        /* 6.Currently enrolled patients that have ever experienced a complication */
        USE openmrs_warehouse;
        ---
        SET @defaultCutOff = '60';
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        SELECT 
            dhi.location,
            gender,
            COUNT(DISTINCT (dhi.patient_id)) AS ncd_count
        FROM
            mw_diabetes_hypertension_initial dhi
                JOIN
            mw_patient mp ON dhi.patient_id = dhi.patient_id
        WHERE
            (cardiovascular_disease IS NOT NULL
                OR retinopathy IS NOT NULL
                OR renal_disease IS NOT NULL
                OR stroke_and_tia IS NOT NULL
                OR peripheral_vascular_disease IS NOT NULL
                OR neuropathy IS NOT NULL
                OR sexual_disorder IS NOT NULL)
                AND diagnosis_hypertension IS NOT NULL
                AND dhi.patient_id IN (SELECT DISTINCT
                    (patient_id)
                FROM
                    (SELECT DISTINCT
                        (mwp.patient_id),
                            opi.identifier,
                            mwp.first_name,
                            mwp.last_name,
                            ops.program,
                            ops.state,
                            ops.start_date,
                            program_state_id,
                            mwp.gender,
                            ops.location,
                            patient_visit.last_appt_date
                    FROM
                        mw_patient mwp
                    JOIN omrs_patient_identifier opi ON mwp.patient_id = opi.patient_id
                    JOIN (SELECT 
                        patient_id,
                            MAX(visit_date) AS visit_date,
                            MAX(next_appointment_date) AS last_appt_date
                    FROM
                        mw_diabetes_hypertension_followup
                    WHERE
                        visit_date <= @endDate
                    GROUP BY patient_id) patient_visit ON patient_visit.patient_id = mwp.patient_id
                    JOIN (SELECT 
                        index_desc,
                            opi.patient_id AS pat,
                            opi.identifier,
                            index_descending.state,
                            index_descending.location,
                            index_descending.program,
                            start_date,
                            program_state_id,
                            end_date
                    FROM
                        (SELECT 
                        @r:=IF(@u = patient_id, @r + 1, 1) index_desc,
                            location,
                            state,
                            program,
                            start_date,
                            end_date,
                            patient_id,
                            program_state_id,
                            @u:=patient_id
                    FROM
                        omrs_program_state, (SELECT @r:=1) AS r, (SELECT @u:=0) AS u
                    WHERE
                        program IN ('Chronic Care Program')
                            AND start_date <= @endDate
                    ORDER BY patient_id DESC , start_date DESC , program_state_id DESC) index_descending
                    JOIN omrs_patient_identifier opi ON index_descending.patient_id = opi.patient_id
                        AND opi.location = index_descending.location
                        AND opi.type = 'Chronic Care Number'
                    WHERE
                        index_desc = 1) ops ON opi.patient_id = ops.pat
                        AND opi.location = ops.location
                    WHERE
                        opi.type = 'Chronic Care Number'
                            AND state IN ('In advanced Care')) x
                WHERE
                    patient_id IN (SELECT 
                            patient_id
                        FROM
                            mw_diabetes_hypertension_followup
                        WHERE
                            FLOOR(DATEDIFF(@endDate, next_appointment_date)) <= @defaultCutOff)
                        AND patient_id IN (SELECT 
                            patient_id
                        FROM
                            mw_diabetes_hypertension_initial
                        WHERE
                            diagnosis_hypertension IS NOT NULL))
        GROUP BY location , gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print(
            "currently_enrolled_patients_that_have_ever_experienced_complication:",
            response,
        )
        write_data(98, response)

    def patients_with_visit_in_last_3_months_BPbelow_140over90(self):
        query = """
        /* 4. Patients with a visit in last 3 months (excluding new patients) that have BP below 140/90 */
        use openmrs_warehouse;
        ---
        SET @startDate = {};
        ---
        SET @endDate = {};
        ---
        call create_chronic_care_outcome(@endDate);
        ---
        SELECT 
            dhf.location, mp.gender, COUNT(*)
        FROM
            mw_diabetes_hypertension_followup dhf
                JOIN
            (SELECT 
                patient_id, MAX(visit_date) AS visit_date
            FROM
                mw_diabetes_hypertension_followup
            WHERE
                visit_date BETWEEN @startDate AND @endDate
            GROUP BY patient_id) dhf1 ON dhf1.patient_id = dhf.patient_id
                AND dhf.visit_date = dhf1.visit_date
                JOIN 
                mw_patient mp ON mp.patient_id = dhf1.patient_id
                JOIN
            chronic_care_last_facility_outcome lfo ON lfo.pat = dhf1.patient_id
                AND dhf.patient_id IN (SELECT 
                    patient_id
                FROM
                    mw_diabetes_hypertension_initial
                WHERE
                    diagnosis_hypertension IS NOT NULL
                        AND visit_date < @startDate)
                AND dhf.bp_stystolic < 140
                AND dhf.bp_diastolic < 90
                AND state = 'In advanced Care'
        GROUP BY dhf.location, mp.gender;
        """
        for end_date in dates:
            if count_placeholders(query) == 2:
                index = dates.index(end_date)
                start_date = start[index]
                formatted_qry = query.format(start_date, end_date)
            else:
                formatted_qry = query.format(end_date)
            qry_lst = formatted_qry.split("---")
            for qry in qry_lst:
                self.cursor_obj.execute(qry)
            response = self.cursor_obj.fetchall()
        print("patients_with_visit_in_last_3_months_BPbelow_140over90:", response)
        write_data(99, response)
